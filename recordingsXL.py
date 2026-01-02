import matplotlib.pyplot as plt
import openpyxl

def recordingsXL(tele_times,who_found_arr,phiI, thetaI, phiI2, thetaI2,rax,decx,A90,tele1,tele2,t_pre_merge,file_name,run,thetaLVK,phiLVK,psiLVK,ILVK,SNRs,track_times,dis,D,SNRs_filtered,t_pre_merge_filtered,A90_filtered ):

    track_time1=track_times[0]
    track_time2 = track_times[1]
    track_time3 = track_times[2]
    time_dif1 = track_time2 - track_time1
    time_dif2 = track_time3 - track_time1
    print(tele1.FOV,'ss')
    print(who_found_arr)
    #plt.plot(2,2)
    #plt.show()
    len_t=len(t_pre_merge_filtered)
    new_line = [
        time_dif1, time_dif2, track_time1, track_time2,track_time3,tele_times[0][0],tele_times[0][1],tele_times[1][0],tele_times[1][1], tele_times[2][0],tele_times[2][1], who_found_arr[0], who_found_arr[1],who_found_arr[2],
        phiI, thetaI, phiI2, thetaI2, rax, decx, A90[0],A90[1],A90[2],A90[3],
        tele1.name, tele2.name, t_pre_merge[0],t_pre_merge[1],t_pre_merge[2],t_pre_merge[3], file_name,
        tele1.max_vel, tele1.acc, tele1.FOV, run,
        thetaLVK, phiLVK, psiLVK, ILVK,SNRs[0],SNRs[1],SNRs[2],SNRs[3],dis,D,len_t
    ]

    # Specify the Excel file name
    excel_file_name = "recordingsP10_13.xlsx"

    try:
        # Try to load existing workbook or create a new one
        wb = openpyxl.load_workbook(excel_file_name)
    except FileNotFoundError:
        # Create a new workbook if the file doesn't exist
        wb = openpyxl.Workbook()

    # Select the active sheet (first sheet by default)
    sheet = wb.active

    # Append the new line as a row in the sheet
    sheet.append(new_line)

    # Save the workbook
    wb.save(excel_file_name)

    # Optionally, print the updated content of the sheet
    for row in sheet.iter_rows(values_only=True):
        print(row)